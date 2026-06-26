import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties

# ---------- AAYNA brand colors ----------
DUSTY_ROSE = "C98B95"
DUSTY_ROSE_LIGHT = "EFD9DC"
CREAM = "FBF7F2"
ANTIQUE_GOLD = "B89B5E"
ESPRESSO = "3B2F2A"
WHITE = "FFFFFF"
SAGE = "8FA98E"
MAUVE = "6B4F4F"
INTEL = "C9A66B"
QTY_COST = "AD8350"
RISK_OPS = "A4626C"
CHECKLIST = "6E5C54"
ACTUAL = "B07B5E"
QC = "6F8F76"
AI_TRACE = "9C8AA5"
GREEN = "C6EFCE"
GREEN_FONT = "1E7B34"
YELLOW = "FFF3C4"
YELLOW_FONT = "9C7A00"
ORANGE = "FCE4C4"
ORANGE_FONT = "B2650A"
RED = "F8CFCB"
RED_FONT = "B23A32"

wb = Workbook()

# =====================================================================
# TAB: SETTINGS  (built first so other tabs can reference named ranges)
# =====================================================================
ws_set = wb.active
ws_set.title = "Settings"
ws_set.sheet_properties.tabColor = ANTIQUE_GOLD

header_fill = PatternFill("solid", fgColor=DUSTY_ROSE)
header_font = Font(bold=True, color=WHITE, size=12)
label_font = Font(bold=True, color=ESPRESSO)
note_font = Font(italic=True, color="6B5B53", size=9)

ws_set["A1"] = "AAYNA Product Scout Lite — Settings"
ws_set["A1"].font = Font(bold=True, size=14, color=ESPRESSO)
ws_set.merge_cells("A1:C1")
ws_set["A2"] = "Tune these numbers once. Every formula in the Product Tracker tab reads from here."
ws_set["A2"].font = note_font
ws_set.merge_cells("A2:D2")

settings_rows = [
    ("Setting", "Value", "Notes"),
    ("USD to BDT exchange rate", 122, "Update when BDT/USD rate changes"),
    ("RMB (CNY) to BDT exchange rate", 17, "Update when BDT/CNY rate changes"),
    ("Default customs/duty %", 0.07, "As decimal, e.g. 0.07 = 7%"),
    ("Default misc fee per unit (BDT)", 15, "Packaging, handling, platform fees per piece"),
    ("Default markup % over landed cost", 1.0, "1.0 = 100% markup (sell at 2x landed cost)"),
    ("Target max selling price (BDT)", 700, "AAYNA's 'mostly under' price ceiling"),
    ("Hard reject price ceiling (BDT)", 1000, "Above the target price but at/under this = 'Price Review' "
        "(negotiable). Above this = automatic 'Reject'."),
    ("Weight: Feminine Score", 0.10, "All 10 weights below must add up to 1.00"),
    ("Weight: Trend Score", 0.10, ""),
    ("Weight: AAYNA Aesthetic Fit Score", 0.15, ""),
    ("Weight: Easy to Style Score", 0.08, ""),
    ("Weight: Lightweight Score", 0.07, ""),
    ("Weight: Reels/Photo Potential Score", 0.10, ""),
    ("Weight: Giftability Score", 0.08, ""),
    ("Weight: Price Fit Score", 0.12, ""),
    ("Weight: Demand Score", 0.12, ""),
    ("Weight: Quality/Risk Score", 0.08, ""),
    ("Buy threshold (score >=)", 75, "Total Product Score is out of 100"),
    ("Maybe threshold (score >=)", 50, "Below this = Reject"),
    ("Low profit margin warning threshold", 0.30, "Auto Flag shows 'Low profit' below this margin, e.g. 0.30 = 30%"),
    ("MOQ warning threshold (units)", 50, "Auto Flag shows 'MOQ too high' above this quantity"),
    ("Monthly Inventory Budget (BDT)", 30000, "Total budget for approved purchases this month — tracked on the Dashboard"),
]
for r, row in enumerate(settings_rows, start=4):
    for c, val in enumerate(row, start=1):
        cell = ws_set.cell(row=r, column=c, value=val)
        if r == 4:
            cell.fill = header_fill
            cell.font = header_font
        else:
            if c == 1:
                cell.font = label_font
            if c == 3:
                cell.font = note_font

ws_set.column_dimensions["A"].width = 36
ws_set.column_dimensions["B"].width = 12
ws_set.column_dimensions["C"].width = 50

# Named ranges for the settings values (column B of each row)
named_map = {
    "USD_BDT": 5,
    "RMB_BDT": 6,
    "CUSTOMS_PCT": 7,
    "MISC_FEE": 8,
    "MARKUP_PCT": 9,
    "MAX_PRICE": 10,
    "HARD_REJECT_PRICE": 11,
    "W_FEMININE": 12,
    "W_TREND": 13,
    "W_AESTHETIC": 14,
    "W_STYLE": 15,
    "W_LIGHT": 16,
    "W_REELS": 17,
    "W_GIFT": 18,
    "W_PRICEFIT": 19,
    "W_DEMAND": 20,
    "W_QUALITY": 21,
    "BUY_THRESHOLD": 22,
    "MAYBE_THRESHOLD": 23,
    "LOW_PROFIT_PCT": 24,
    "MOQ_WARNING": 25,
    "MONTHLY_BUDGET": 26,
}
for name, row in named_map.items():
    ref = f"Settings!$B${row}"
    wb.defined_names[name] = DefinedName(name, attr_text=ref)

# =====================================================================
# TAB: DROPDOWNS  (lists used for data validation everywhere)
# =====================================================================
ws_lists = wb.create_sheet("Dropdown Lists")
ws_lists.sheet_properties.tabColor = ANTIQUE_GOLD

lists = {
    "A": ("Source Platform", ["SkyBuyBD", "AliExpress", "Yiwugo", "1688", "Local Wholesale", "Other"]),
    "B": ("Category", ["Earrings", "Necklace", "Bracelet", "Ring", "Hair Accessory", "Bag", "Belt",
                        "Sunglasses", "Watch", "Scarf", "Hijab Accessory", "Gift Set", "Other"]),
    "C": ("Source Currency", ["BDT", "USD", "RMB"]),
    "D": ("Decision", ["Buy", "Maybe", "Price Review", "Reject"]),
    "E": ("Approval Status", ["Pending", "Approved", "Rejected", "On Hold"]),
    "F": ("Score (1-5)", [1, 2, 3, 4, 5]),
    "G": ("Yes/No", ["Yes", "No"]),
    "H": ("Risk Level", ["Low", "Medium", "High"]),
    "I": ("Packaging Difficulty", ["Easy", "Medium", "Hard"]),
    "J": ("Sourcing Status", ["Not Started", "Sourcing", "Sample Ordered", "Sample Received",
                               "Approved - Not Ordered", "Ordered", "In Transit", "Arrived", "Live on Website"]),
    "K": ("QC Status", ["Not Arrived", "QC Pending", "QC Passed", "Minor Defect", "QC Failed",
                         "Discount Sell", "Return/Reject"]),
    "L": ("AI Tool", ["Claude", "ChatGPT", "Gemini", "Other"]),
}
for col, (title, values) in lists.items():
    ws_lists[f"{col}1"] = title
    ws_lists[f"{col}1"].fill = header_fill
    ws_lists[f"{col}1"].font = header_font
    for i, v in enumerate(values, start=2):
        ws_lists[f"{col}{i}"] = v
    ws_lists.column_dimensions[col].width = 18

# Named ranges for each list (fixed length covers max list size)
wb.defined_names["LIST_PLATFORM"] = DefinedName("LIST_PLATFORM", attr_text="'Dropdown Lists'!$A$2:$A$7")
wb.defined_names["LIST_CATEGORY"] = DefinedName("LIST_CATEGORY", attr_text="'Dropdown Lists'!$B$2:$B$14")
wb.defined_names["LIST_CURRENCY"] = DefinedName("LIST_CURRENCY", attr_text="'Dropdown Lists'!$C$2:$C$4")
wb.defined_names["LIST_DECISION"] = DefinedName("LIST_DECISION", attr_text="'Dropdown Lists'!$D$2:$D$5")
wb.defined_names["LIST_APPROVAL"] = DefinedName("LIST_APPROVAL", attr_text="'Dropdown Lists'!$E$2:$E$5")
wb.defined_names["LIST_SCORE"] = DefinedName("LIST_SCORE", attr_text="'Dropdown Lists'!$F$2:$F$6")
wb.defined_names["LIST_YESNO"] = DefinedName("LIST_YESNO", attr_text="'Dropdown Lists'!$G$2:$G$3")
wb.defined_names["LIST_LEVEL"] = DefinedName("LIST_LEVEL", attr_text="'Dropdown Lists'!$H$2:$H$4")
wb.defined_names["LIST_PACKAGING"] = DefinedName("LIST_PACKAGING", attr_text="'Dropdown Lists'!$I$2:$I$4")
wb.defined_names["LIST_SOURCING_STATUS"] = DefinedName("LIST_SOURCING_STATUS", attr_text="'Dropdown Lists'!$J$2:$J$10")
wb.defined_names["LIST_QC_STATUS"] = DefinedName("LIST_QC_STATUS", attr_text="'Dropdown Lists'!$K$2:$K$8")
wb.defined_names["LIST_AI_TOOL"] = DefinedName("LIST_AI_TOOL", attr_text="'Dropdown Lists'!$L$2:$L$5")

# Fixed category -> SKU prefix lookup table (explicit, not derived from the
# category name) so SKUs stay stable even if a category is renamed slightly.
category_prefixes = [
    ("Earrings", "EAR"),
    ("Necklace", "NEC"),
    ("Bracelet", "BRC"),
    ("Ring", "RNG"),
    ("Hair Accessory", "HAR"),
    ("Bag", "BAG"),
    ("Belt", "BLT"),
    ("Sunglasses", "SUN"),
    ("Watch", "WCH"),
    ("Scarf", "SCF"),
    ("Hijab Accessory", "HIJ"),
    ("Gift Set", "GFT"),
    ("Other", "OTH"),
]
ws_lists["M1"] = "Category"
ws_lists["N1"] = "SKU Prefix"
ws_lists["M1"].fill = header_fill
ws_lists["N1"].fill = header_fill
ws_lists["M1"].font = header_font
ws_lists["N1"].font = header_font
for i, (cat, prefix) in enumerate(category_prefixes, start=2):
    ws_lists[f"M{i}"] = cat
    ws_lists[f"N{i}"] = prefix
ws_lists.column_dimensions["M"].width = 18
ws_lists.column_dimensions["N"].width = 12
wb.defined_names["CATEGORY_PREFIX_TABLE"] = DefinedName(
    "CATEGORY_PREFIX_TABLE", attr_text=f"'Dropdown Lists'!$M$2:$N${1 + len(category_prefixes)}")

# =====================================================================
# TAB: PRODUCT TRACKER (the main sheet)
# =====================================================================
ws = wb.create_sheet("Product Tracker", 0)
ws.sheet_properties.tabColor = DUSTY_ROSE

columns = [
    # (header, width, group_color) -- 35 visible columns, A..AI
    ("SKU", 12, DUSTY_ROSE),
    ("Date Added", 12, DUSTY_ROSE),
    ("Product Name", 26, DUSTY_ROSE),
    ("Product Image/Link", 22, DUSTY_ROSE),
    ("Category", 16, DUSTY_ROSE),
    ("Source Platform", 14, DUSTY_ROSE),
    ("Source URL", 22, DUSTY_ROSE),
    ("Supplier Name", 18, DUSTY_ROSE),

    ("Unit Cost (Source Currency)", 14, ANTIQUE_GOLD),
    ("Source Currency", 12, ANTIQUE_GOLD),
    ("Unit Cost (BDT)", 13, ANTIQUE_GOLD),
    ("Shipping Cost/Unit (BDT)", 14, ANTIQUE_GOLD),
    ("Customs/Duty %", 12, ANTIQUE_GOLD),
    ("Misc Fees/Unit (BDT)", 13, ANTIQUE_GOLD),
    ("Estimated Landed Cost (BDT)", 15, ANTIQUE_GOLD),

    ("Suggested Selling Price (BDT)", 16, SAGE),
    ("Expected Profit/Unit (BDT)", 14, SAGE),
    ("Profit Margin %", 12, SAGE),

    ("Feminine Score (1-5)", 12, ESPRESSO),
    ("Trend Score (1-5)", 11, ESPRESSO),
    ("AAYNA Aesthetic Fit Score (1-5)", 14, ESPRESSO),
    ("Easy to Style Score (1-5)", 12, ESPRESSO),
    ("Lightweight Score (1-5)", 12, ESPRESSO),
    ("Reels/Photo Potential Score (1-5)", 14, ESPRESSO),
    ("Giftability Score (1-5)", 12, ESPRESSO),
    ("Price Fit Score (1-5)", 12, ESPRESSO),
    ("Demand Score (1-5)", 12, ESPRESSO),
    ("Quality/Risk Score (1-5)", 12, ESPRESSO),

    ("Total Product Score (/100)", 13, ESPRESSO),
    ("Decision", 13, ESPRESSO),
    ("Reason", 28, ESPRESSO),

    ("Content/Reel Idea", 30, MAUVE),
    ("Approval Status", 13, MAUVE),
    ("Approved By", 14, MAUVE),
    ("Date Decided", 13, MAUVE),
]

# Hidden helper columns (AJ..AN) -- rank keys used by the Dashboard tab to find
# Top-N products without relying on QUERY/FILTER/SORT (not available in every
# spreadsheet app). Each key subtracts ROW()*0.0000001 so every key is unique,
# which keeps LARGE()+MATCH() lookups on the Dashboard from colliding on ties.
helper_columns = [
    "Rank Key: Overall (hidden)",
    "Rank Key: Under 700 BDT (hidden)",
    "Rank Key: Reels/Photo (hidden)",
    "Rank Key: Giftability (hidden)",
    "Rank Key: Needs Review (hidden)",
]

n_data_rows = 498  # rows 3..500
header_row = 2

ws["A1"] = "AAYNA Product Scout Lite — Product Tracker"
ws["A1"].font = Font(bold=True, size=14, color=ESPRESSO)
ws.merge_cells("A1:H1")
ws["I1"] = "Reflect your everyday style — sourcing decisions, made together."
ws["I1"].font = Font(italic=True, color="6B5B53", size=10)
ws.merge_cells("I1:R1")

for idx, (name, width, color) in enumerate(columns, start=1):
    col_letter = get_column_letter(idx)
    cell = ws.cell(row=header_row, column=idx, value=name)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.column_dimensions[col_letter].width = width

helper_start_idx = len(columns) + 1  # AJ
for offset, name in enumerate(helper_columns):
    idx = helper_start_idx + offset
    col_letter = get_column_letter(idx)
    cell = ws.cell(row=header_row, column=idx, value=name)
    cell.fill = PatternFill("solid", fgColor="BFBFBF")
    cell.font = Font(bold=True, color=WHITE, size=9)
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.column_dimensions[col_letter].width = 10
    ws.column_dimensions[col_letter].hidden = True

# Sourcing & operations columns (AO..BL) -- appended after the hidden rank-key
# columns so none of the existing column letters above (used throughout the
# Decision/score formulas) have to move.
new_columns = [
    ("Supplier Rating", 12, INTEL),
    ("Product Rating", 12, INTEL),
    ("Sold Count / Order Count", 14, INTEL),
    ("Review Count", 12, INTEL),
    ("Real Review Photos?", 13, INTEL),

    ("MOQ", 10, QTY_COST),
    ("Recommended Test Quantity", 14, QTY_COST),
    ("Final Approved Quantity", 14, QTY_COST),
    ("Total Purchase Cost (BDT)", 15, QTY_COST),

    ("Weight / Size", 18, RISK_OPS),
    ("Fragility Level", 12, RISK_OPS),
    ("Packaging Difficulty", 13, RISK_OPS),
    ("Courier Risk", 12, RISK_OPS),
    ("Duplicate Found on BD Market?", 14, RISK_OPS),
    ("Competitor Price (BDT)", 13, RISK_OPS),
    ("Market Saturation Level", 13, RISK_OPS),

    ("Auto Flag / Warning", 17, ESPRESSO),
    ("Sourcing Status", 16, ESPRESSO),

    ("Product Name Finalized?", 13, CHECKLIST),
    ("SKU Finalized?", 13, CHECKLIST),
    ("Photos Ready?", 13, CHECKLIST),
    ("Description Ready?", 13, CHECKLIST),
    ("Price Approved?", 13, CHECKLIST),
    ("Ready for Website Upload?", 15, CHECKLIST),
]
new_start_idx = helper_start_idx + len(helper_columns)  # AO
for offset, (name, width, color) in enumerate(new_columns):
    idx = new_start_idx + offset
    col_letter = get_column_letter(idx)
    cell = ws.cell(row=header_row, column=idx, value=name)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.column_dimensions[col_letter].width = width

# Actual cost/QC/AI-trace columns (BM..CB) -- appended after the sourcing &
# operations block for the same reason: nothing above has to move.
new_columns2 = [
    ("Actual Product Cost (BDT)", 15, ACTUAL),
    ("Actual Shipping Cost (BDT)", 15, ACTUAL),
    ("Actual Landed Cost (BDT)", 14, ACTUAL),
    ("Actual Selling Price (BDT)", 14, ACTUAL),
    ("Actual Profit (BDT)", 13, ACTUAL),
    ("Actual Profit Margin", 13, ACTUAL),

    ("Arrival Date", 13, QC),
    ("Quantity Received", 13, QC),
    ("Defect Count", 12, QC),
    ("QC Status", 13, QC),
    ("QC Notes", 26, QC),
    ("Final Stock Accepted", 14, QC),

    ("AI Tool Used", 13, AI_TRACE),
    ("AI Score Date", 13, AI_TRACE),
    ("Scored By", 13, AI_TRACE),
    ("Manual Score Adjusted?", 15, AI_TRACE),
]
new2_start_idx = new_start_idx + len(new_columns)  # BM
for offset, (name, width, color) in enumerate(new_columns2):
    idx = new2_start_idx + offset
    col_letter = get_column_letter(idx)
    cell = ws.cell(row=header_row, column=idx, value=name)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.column_dimensions[col_letter].width = width

visible_col_indices = (list(range(1, len(columns) + 1))
                        + list(range(new_start_idx, new_start_idx + len(new_columns)))
                        + list(range(new2_start_idx, new2_start_idx + len(new_columns2))))

ws.row_dimensions[header_row].height = 42
ws.freeze_panes = "C3"

thin = Side(style="thin", color="D9CDC3")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

first_data_row = header_row + 1
last_data_row = header_row + n_data_rows

ws.auto_filter.ref = f"A{header_row}:{get_column_letter(new2_start_idx + len(new_columns2) - 1)}{last_data_row}"

for r in range(first_data_row, last_data_row + 1):
    # SKU (A): AYN-<fixed category prefix>-<4-digit row number>, e.g. AYN-EAR-0001.
    # Prefix comes from the CATEGORY_PREFIX_TABLE lookup, not from the category
    # text itself, so it stays correct even for categories that share a first
    # 3 letters (e.g. Belt vs Bracelet).
    ws[f"A{r}"] = (
        f'=IF(C{r}="","","AYN-"&IFERROR(VLOOKUP(E{r},CATEGORY_PREFIX_TABLE,2,FALSE),"OTH")'
        f'&"-"&TEXT(ROW()-{header_row},"0000"))'
    )
    # Unit Cost (BDT) (K)
    ws[f"K{r}"] = f'=IF(I{r}="","",IF(J{r}="BDT",I{r},IF(J{r}="USD",I{r}*USD_BDT,IF(J{r}="RMB",I{r}*RMB_BDT,""))))'
    # Estimated Landed Cost (O) = unit cost bdt + shipping + customs%*unit cost bdt + misc fees
    # (M and N fall back to Settings defaults via the IF()s below when left blank)
    ws[f"O{r}"] = f'=IF(K{r}="","",K{r}+IFERROR(L{r},0)+(K{r}*IF(M{r}="",CUSTOMS_PCT,M{r}))+IF(N{r}="",MISC_FEE,N{r}))'
    # Suggested Selling Price (P)
    ws[f"P{r}"] = f'=IF(O{r}="","",ROUND(O{r}*(1+MARKUP_PCT),-1)-1)'
    # Expected Profit/Unit (Q)
    ws[f"Q{r}"] = f'=IF(P{r}="","",P{r}-O{r})'
    # Profit Margin % (R)
    ws[f"R{r}"] = f'=IF(OR(P{r}="",P{r}=0),"",Q{r}/P{r})'
    # Total Product Score (AC), weighted across all 10 criteria, out of 100
    ws[f"AC{r}"] = (
        f'=IF(COUNT(S{r}:AB{r})<10,"",'
        f'(S{r}*W_FEMININE+T{r}*W_TREND+U{r}*W_AESTHETIC+V{r}*W_STYLE+W{r}*W_LIGHT+'
        f'X{r}*W_REELS+Y{r}*W_GIFT+Z{r}*W_PRICEFIT+AA{r}*W_DEMAND+AB{r}*W_QUALITY)/5*100)'
    )
    # Decision (AD)
    # Price-based hard rules run first and apply even if scoring isn't finished yet:
    #   price > HARD_REJECT_PRICE  -> Reject outright
    #   MAX_PRICE < price <= HARD_REJECT_PRICE -> Price Review (negotiable, not dead)
    # Once price clears the bar, a low Quality/Risk or Aesthetic Fit score (<=2)
    # caps a would-be "Buy" down to "Maybe" instead of forcing a flat Reject.
    ws[f"AD{r}"] = (
        f'=IF(P{r}="","",'
        f'IF(P{r}>HARD_REJECT_PRICE,"Reject",'
        f'IF(P{r}>MAX_PRICE,"Price Review",'
        f'IF(AC{r}="","",'
        f'IF(AND(AC{r}>=BUY_THRESHOLD,OR(AB{r}<=2,U{r}<=2)),"Maybe",'
        f'IF(AC{r}>=BUY_THRESHOLD,"Buy",'
        f'IF(AC{r}>=MAYBE_THRESHOLD,"Maybe","Reject")))))))'
    )
    # ---- Hidden rank-key helper columns for the Dashboard tab ----
    ws[f"AJ{r}"] = f'=IF(AC{r}="",-9999-ROW()*0.0000001,AC{r}-ROW()*0.0000001)'
    ws[f"AK{r}"] = f'=IF(OR(AC{r}="",P{r}="",P{r}>MAX_PRICE),-9999-ROW()*0.0000001,AC{r}-ROW()*0.0000001)'
    ws[f"AL{r}"] = f'=IF(AC{r}="",-9999-ROW()*0.0000001,X{r}*1000+AC{r}-ROW()*0.0000001)'
    ws[f"AM{r}"] = f'=IF(AC{r}="",-9999-ROW()*0.0000001,Y{r}*1000+AC{r}-ROW()*0.0000001)'
    ws[f"AN{r}"] = (
        f'=IF(AND(OR(AD{r}="Maybe",AD{r}="Price Review"),AG{r}<>"Approved",AG{r}<>"Rejected"),'
        f'IF(AC{r}="",0,AC{r})-ROW()*0.0000001,-9999-ROW()*0.0000001)'
    )

    # ---- Sourcing & operations formulas ----
    # Total Purchase Cost (AW) = Estimated Landed Cost x Final Approved Quantity
    ws[f"AW{r}"] = f'=IF(OR(O{r}="",AV{r}=""),"",O{r}*AV{r})'
    # Auto Flag / Warning (BE): first matching problem wins, in priority order.
    # ISNUMBER() guards every numeric comparison so a still-blank score/price/MOQ
    # cell can never be misread as 0 (which would falsely look "too low").
    ws[f"BE{r}"] = (
        f'=IF(C{r}="","",'
        f'IF(AND(ISNUMBER(P{r}),P{r}>MAX_PRICE),"Price too high",'
        f'IF(AND(ISNUMBER(AB{r}),AB{r}<=2),"Quality risk",'
        f'IF(AND(ISNUMBER(U{r}),U{r}<=2),"Aesthetic mismatch",'
        f'IF(AY{r}="High","Too fragile",'
        f'IF(AND(ISNUMBER(R{r}),R{r}<LOW_PROFIT_PCT),"Low profit",'
        f'IF(AND(ISNUMBER(AT{r}),AT{r}>MOQ_WARNING),"MOQ too high",'
        f'IF(OR(AD{r}="Maybe",AD{r}="Price Review"),"Needs partner review",'
        f'IF(AD{r}="Buy","Good candidate","")))))))))'
    )
    # Ready for Website Upload? (BL) -- Yes only once every checklist item is Yes
    ws[f"BL{r}"] = (
        f'=IF(C{r}="","",IF(AND(BG{r}="Yes",BH{r}="Yes",BI{r}="Yes",BJ{r}="Yes",BK{r}="Yes"),"Yes","No"))'
    )

    # ---- Actual cost & profit (filled in once the order has actually shipped) ----
    # Actual Landed Cost (BO) = Actual Product Cost + Actual Shipping Cost
    ws[f"BO{r}"] = f'=IF(OR(BM{r}="",BN{r}=""),"",BM{r}+BN{r})'
    # Actual Profit (BQ) = Actual Selling Price - Actual Landed Cost
    ws[f"BQ{r}"] = f'=IF(OR(BP{r}="",BO{r}=""),"",BP{r}-BO{r})'
    # Actual Profit Margin (BR)
    ws[f"BR{r}"] = f'=IF(OR(BP{r}="",BP{r}=0,BQ{r}=""),"",BQ{r}/BP{r})'

    # ---- Quality check after arrival ----
    # Final Stock Accepted (BX) = Quantity Received - Defect Count
    ws[f"BX{r}"] = f'=IF(OR(BT{r}="",BU{r}=""),"",BT{r}-BU{r})'

    for c in visible_col_indices:
        ws.cell(row=r, column=c).border = border
        ws.cell(row=r, column=c).alignment = Alignment(vertical="center")

# Number formats
for r in range(first_data_row, last_data_row + 1):
    ws[f"B{r}"].number_format = "yyyy-mm-dd"
    for col in ["K", "L", "N", "O", "P", "Q"]:
        ws[f"{col}{r}"].number_format = '#,##0 "BDT"'
    ws[f"M{r}"].number_format = "0%"
    ws[f"R{r}"].number_format = "0%"
    ws[f"AC{r}"].number_format = "0"
    ws[f"AI{r}"].number_format = "yyyy-mm-dd"
    for col in ["AO", "AP"]:
        ws[f"{col}{r}"].number_format = "0.0"
    for col in ["AQ", "AR", "AT", "AU", "AV"]:
        ws[f"{col}{r}"].number_format = "#,##0"
    for col in ["AW", "BC"]:
        ws[f"{col}{r}"].number_format = '#,##0 "BDT"'
    for col in ["BM", "BN", "BO", "BP", "BQ"]:
        ws[f"{col}{r}"].number_format = '#,##0 "BDT"'
    ws[f"BR{r}"].number_format = "0%"
    ws[f"BS{r}"].number_format = "yyyy-mm-dd"
    for col in ["BT", "BU", "BX"]:
        ws[f"{col}{r}"].number_format = "#,##0"
    ws[f"BZ{r}"].number_format = "yyyy-mm-dd"

# Light banding fill for readability
band_fill = PatternFill("solid", fgColor=CREAM)
for r in range(first_data_row, last_data_row + 1):
    if (r - first_data_row) % 2 == 1:
        for c in visible_col_indices:
            ws.cell(row=r, column=c).fill = band_fill

# ---------------- Data validation dropdowns ----------------
def add_dv(formula1, col_letter, tooltip=""):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True, showDropDown=False)
    dv.error = "Please choose a value from the list."
    dv.errorTitle = "Invalid entry"
    dv.prompt = tooltip
    dv.promptTitle = "Choose from list"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_data_row}:{col_letter}{last_data_row}")
    return dv

add_dv("LIST_PLATFORM", "F", "Where this product is sourced from")
add_dv("LIST_CATEGORY", "E", "Product category")
add_dv("LIST_CURRENCY", "J", "Currency the unit cost is quoted in")
add_dv("LIST_APPROVAL", "AG", "Partner approval status")
score_cols = ["S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB"]
for col in score_cols:
    add_dv("LIST_SCORE", col, "Score 1 (worst) to 5 (best)")

yesno_cols = ["AS", "BB", "BG", "BH", "BI", "BJ", "BK"]
for col in yesno_cols:
    add_dv("LIST_YESNO", col, "Yes or No")
for col in ["AY", "BA", "BD"]:
    add_dv("LIST_LEVEL", col, "Low, Medium, or High")
add_dv("LIST_PACKAGING", "AZ", "Easy, Medium, or Hard")
add_dv("LIST_SOURCING_STATUS", "BF", "Where this product is in the sourcing pipeline")
add_dv("LIST_QC_STATUS", "BV", "Quality check result after the product arrives")
add_dv("LIST_AI_TOOL", "BY", "Which AI tool gave the first-opinion scores")
add_dv("LIST_YESNO", "CB", "Did a human change any AI-suggested score?")

# ---------------- Conditional formatting on Decision (AD) ----------------
rng = f"AD{first_data_row}:AD{last_data_row}"
ws.conditional_formatting.add(
    rng, CellIsRule(operator="equal", formula=['"Buy"'],
                     fill=PatternFill("solid", fgColor=GREEN), font=Font(color=GREEN_FONT, bold=True)))
ws.conditional_formatting.add(
    rng, CellIsRule(operator="equal", formula=['"Maybe"'],
                     fill=PatternFill("solid", fgColor=YELLOW), font=Font(color=YELLOW_FONT, bold=True)))
ws.conditional_formatting.add(
    rng, CellIsRule(operator="equal", formula=['"Price Review"'],
                     fill=PatternFill("solid", fgColor=ORANGE), font=Font(color=ORANGE_FONT, bold=True)))
ws.conditional_formatting.add(
    rng, CellIsRule(operator="equal", formula=['"Reject"'],
                     fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_FONT, bold=True)))

# Highlight Suggested Selling Price if it crosses either price ceiling
price_rng = f"P{first_data_row}:P{last_data_row}"
ws.conditional_formatting.add(
    price_rng,
    FormulaRule(formula=[f"P{first_data_row}>HARD_REJECT_PRICE"], font=Font(color=RED_FONT, bold=True)))
ws.conditional_formatting.add(
    price_rng,
    FormulaRule(formula=[f"AND(P{first_data_row}>MAX_PRICE,P{first_data_row}<=HARD_REJECT_PRICE)"],
                font=Font(color=ORANGE_FONT, bold=True)))

# Warn on low Aesthetic Fit (U) and Quality/Risk (AB) scores -- these are the
# two scores that cap a "Buy" down to "Maybe" per the hard rejection rules.
for col in ["U", "AB"]:
    warn_rng = f"{col}{first_data_row}:{col}{last_data_row}"
    ws.conditional_formatting.add(
        warn_rng, CellIsRule(operator="lessThanOrEqual", formula=["2"],
                              fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_FONT, bold=True)))

# Approval status colors
appr_rng = f"AG{first_data_row}:AG{last_data_row}"
ws.conditional_formatting.add(appr_rng, CellIsRule(operator="equal", formula=['"Approved"'],
                     fill=PatternFill("solid", fgColor=GREEN)))
ws.conditional_formatting.add(appr_rng, CellIsRule(operator="equal", formula=['"Rejected"'],
                     fill=PatternFill("solid", fgColor=RED)))
ws.conditional_formatting.add(appr_rng, CellIsRule(operator="equal", formula=['"Pending"'],
                     fill=PatternFill("solid", fgColor=YELLOW)))

# Auto Flag / Warning (BE) colors -- green for a clean pass, yellow for a
# routine partner-review nudge, red for any of the specific risk flags.
flag_rng = f"BE{first_data_row}:BE{last_data_row}"
ws.conditional_formatting.add(flag_rng, CellIsRule(operator="equal", formula=['"Good candidate"'],
                     fill=PatternFill("solid", fgColor=GREEN), font=Font(color=GREEN_FONT, bold=True)))
ws.conditional_formatting.add(flag_rng, CellIsRule(operator="equal", formula=['"Needs partner review"'],
                     fill=PatternFill("solid", fgColor=YELLOW), font=Font(color=YELLOW_FONT, bold=True)))
ws.conditional_formatting.add(
    flag_rng,
    FormulaRule(formula=[f'AND(BE{first_data_row}<>"",BE{first_data_row}<>"Good candidate",'
                          f'BE{first_data_row}<>"Needs partner review")'],
                fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_FONT, bold=True)))

# Ready for Website Upload (BL) colors
upload_rng = f"BL{first_data_row}:BL{last_data_row}"
ws.conditional_formatting.add(upload_rng, CellIsRule(operator="equal", formula=['"Yes"'],
                     fill=PatternFill("solid", fgColor=GREEN), font=Font(color=GREEN_FONT, bold=True)))
ws.conditional_formatting.add(upload_rng, CellIsRule(operator="equal", formula=['"No"'],
                     fill=PatternFill("solid", fgColor=YELLOW)))

# Risk-level colors -- Fragility Level (AY), Courier Risk (BA), Market
# Saturation Level (BD) all share the same Low/Medium/High scale.
for col in ["AY", "BA", "BD"]:
    risk_rng = f"{col}{first_data_row}:{col}{last_data_row}"
    ws.conditional_formatting.add(risk_rng, CellIsRule(operator="equal", formula=['"High"'],
                         fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_FONT, bold=True)))
    ws.conditional_formatting.add(risk_rng, CellIsRule(operator="equal", formula=['"Medium"'],
                         fill=PatternFill("solid", fgColor=YELLOW)))
    ws.conditional_formatting.add(risk_rng, CellIsRule(operator="equal", formula=['"Low"'],
                         fill=PatternFill("solid", fgColor=GREEN)))

# QC Status (BV) colors
qc_rng = f"BV{first_data_row}:BV{last_data_row}"
ws.conditional_formatting.add(qc_rng, CellIsRule(operator="equal", formula=['"QC Passed"'],
                     fill=PatternFill("solid", fgColor=GREEN), font=Font(color=GREEN_FONT, bold=True)))
ws.conditional_formatting.add(qc_rng, CellIsRule(operator="equal", formula=['"QC Pending"'],
                     fill=PatternFill("solid", fgColor=YELLOW), font=Font(color=YELLOW_FONT, bold=True)))
for val in ['"Minor Defect"', '"Discount Sell"']:
    ws.conditional_formatting.add(qc_rng, CellIsRule(operator="equal", formula=[val],
                         fill=PatternFill("solid", fgColor=ORANGE), font=Font(color=ORANGE_FONT, bold=True)))
for val in ['"QC Failed"', '"Return/Reject"']:
    ws.conditional_formatting.add(qc_rng, CellIsRule(operator="equal", formula=[val],
                         fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_FONT, bold=True)))

# ---------------- Sample example rows (first 5 data rows) ----------------
# Each row is built to land on a different branch of the Decision formula so
# partners can see exactly how the hard rejection rules behave.
sample_buy = {
    "B": "2026-06-20", "C": "Antique Gold Hoop Earrings", "D": "drive.google.com/sample1",
    "E": "Earrings", "F": "AliExpress", "G": "aliexpress.com/item/sample1", "H": "Lin Wei Trading",
    "I": 1.8, "J": "USD", "L": 25, "N": 15,
    "S": 5, "T": 5, "U": 5, "V": 4, "W": 5, "X": 5, "Y": 4, "Z": 5, "AA": 4, "AB": 5,
    "AE": "Strong rose-gold match, trending on TikTok, lightweight and giftable.",
    "AF": "Reel: 'Get Ready With Me' earring swap, 3 looks in 15 sec.",
    "AG": "Approved", "AH": "Nadia", "AI": "2026-06-21",
    "AO": 4.8, "AP": 4.7, "AQ": 1200, "AR": 340, "AS": "Yes",
    "AT": 20, "AU": 30, "AV": 50,
    "AX": "15g, 3 x 4 cm", "AY": "Low", "AZ": "Easy", "BA": "Low", "BB": "No", "BC": 650, "BD": "Low",
    "BF": "Approved - Not Ordered",
    "BG": "Yes", "BH": "Yes", "BI": "Yes", "BJ": "Yes", "BK": "Yes",
    "BV": "Not Arrived",
    "BY": "Claude", "BZ": "2026-06-20", "CA": "Nadia", "CB": "No",
}
sample_maybe_quality_cap = {
    "B": "2026-06-21", "C": "Beaded Choker Necklace", "D": "drive.google.com/sample2",
    "E": "Necklace", "F": "SkyBuyBD", "G": "skybuybd.com/item/sample2", "H": "Skybuy Supplier 14",
    "I": 60, "J": "BDT", "L": 10, "N": 12,
    "S": 5, "T": 4, "U": 5, "V": 4, "W": 4, "X": 4, "Y": 4, "Z": 5, "AA": 4, "AB": 2,
    "AE": "Gorgeous and on-trend, but beads chip easily in transit — quality/return risk caps this at Maybe.",
    "AF": "Try-on reel layering 3 chokers.",
    "AG": "Pending",
    "AO": 4.3, "AP": 4.1, "AQ": 650, "AR": 180, "AS": "Yes",
    "AT": 30, "AU": 20,
    "AX": "25g, one size", "AY": "Medium", "AZ": "Medium", "BA": "Medium", "BB": "Yes", "BC": 80, "BD": "Medium",
    "BF": "Sample Ordered",
    "BV": "Not Arrived",
    "BY": "ChatGPT", "BZ": "2026-06-21", "CA": "Imran", "CB": "Yes",
}
sample_reject_low_score = {
    "B": "2026-06-21", "C": "Plain Plastic Hair Clip", "D": "drive.google.com/sample3",
    "E": "Hair Accessory", "F": "1688", "G": "1688.com/item/sample3", "H": "Generic Factory 8",
    "I": 25, "J": "BDT", "L": 8, "N": 10,
    "S": 2, "T": 2, "U": 2, "V": 3, "W": 3, "X": 2, "Y": 2, "Z": 4, "AA": 2, "AB": 3,
    "AE": "Generic look, low trend/demand/aesthetic fit — total score too low to justify stocking.",
    "AF": "",
    "AG": "Rejected",
    "AO": 3.2, "AP": 3.0, "AQ": 2200, "AR": 90, "AS": "No",
    "AT": 80,
    "AX": "5g, 6 cm clip", "AY": "Low", "AZ": "Easy", "BA": "Low", "BB": "Yes", "BC": 15, "BD": "High",
    "BF": "Not Started",
    "BV": "Not Arrived",
    "BY": "Claude", "BZ": "2026-06-21", "CA": "Nadia", "CB": "No",
}
sample_price_review = {
    "B": "2026-06-22", "C": "Embroidered Tote Bag", "D": "drive.google.com/sample4",
    "E": "Bag", "F": "AliExpress", "G": "aliexpress.com/item/sample4", "H": "Lin Wei Trading",
    "I": 300, "J": "BDT", "L": 40, "N": 15,
    "S": 4, "T": 4, "U": 4, "V": 3, "W": 3, "X": 3, "Y": 4, "Z": 2, "AA": 4, "AB": 4,
    "AE": "Lovely design, but landed cost pushes selling price over 700 — needs supplier price negotiation.",
    "AF": "Flatlay reel with 3 outfit pairings.",
    "AG": "Pending",
    "AO": 4.5, "AP": 4.2, "AQ": 300, "AR": 75, "AS": "Yes",
    "AT": 40, "AU": 15,
    "AX": "320g, 30 x 35 cm", "AY": "Low", "AZ": "Medium", "BA": "Medium", "BB": "No", "BC": 720, "BD": "Low",
    "BF": "Sourcing",
    "BG": "Yes", "BH": "Yes",
    "BV": "Not Arrived",
    "BY": "Gemini", "BZ": "2026-06-22", "CA": "Imran", "CB": "No",
}
sample_hard_reject_price = {
    "B": "2026-06-22", "C": "Designer-Inspired Sunglasses", "D": "drive.google.com/sample5",
    "E": "Sunglasses", "F": "AliExpress", "G": "aliexpress.com/item/sample5", "H": "Guangzhou Eyewear Co",
    "I": 15, "J": "USD", "L": 50, "N": 15,
    "AE": "Cost alone puts this far above our price ceiling even before scoring — auto-rejected.",
    "AF": "",
    "AG": "Rejected",
    "AO": 3.8, "AP": 3.5, "AQ": 900, "AR": 210, "AS": "No",
    "AT": 60,
    "AX": "28g, with case", "AY": "High", "AZ": "Hard", "BA": "High", "BB": "Yes", "BC": 450, "BD": "High",
    "BF": "Not Started",
    "BV": "Not Arrived",
    "BY": "ChatGPT", "BZ": "2026-06-22", "CA": "Nadia", "CB": "No",
}
samples = [sample_buy, sample_maybe_quality_cap, sample_reject_low_score,
           sample_price_review, sample_hard_reject_price]
for sample, row in zip(samples, range(first_data_row, first_data_row + len(samples))):
    for col, val in sample.items():
        ws[f"{col}{row}"] = val

# =====================================================================
# TAB: DASHBOARD
# =====================================================================
# Visual-only redesign: this block uses its own AAYNA brand palette (DASH_*)
# instead of the shared color constants above, so the Product Tracker,
# Settings, Dropdown Lists and Claude Scoring Prompt tabs are unaffected.
DASH_ROSE = "9A4F5F"
DASH_CREAM = "FFF8F2"
DASH_MIST = "F7E7E9"
DASH_GOLD = "C6A15B"
DASH_ESPRESSO = "2F2623"
DASH_TAUPE = "7B6A63"
DASH_BORDER = "EADDD4"
DASH_WHITE = "FFFFFF"
MONEY_FMT = '৳#,##0'

ws_dash = wb.create_sheet("Dashboard")
ws_dash.sheet_properties.tabColor = SAGE
ws_dash.sheet_view.showGridLines = False

PT = "'Product Tracker'"
pt_first, pt_last = first_data_row, last_data_row

for col, width in zip("ABCDEFGHIJKL", [6, 11, 11, 11, 11, 9, 9, 8, 9, 9, 9, 9]):
    ws_dash.column_dimensions[col].width = width

beige_thin = Side(style="thin", color=DASH_BORDER)
header_bottom_border = Border(bottom=Side(style="medium", color=DASH_GOLD))
row_border = Border(left=beige_thin, right=beige_thin, top=beige_thin, bottom=beige_thin)

# ---- Hidden Dashboard-local rank-key helpers (row-aligned with Product
# Tracker rows 3-500) so new ranked sections can filter by Decision/Website
# Ready/High Risk without adding or editing anything on Product Tracker. ----
for r in range(pt_first, pt_last + 1):
    ws_dash[f"R{r}"] = f'=IF({PT}!AD{r}="Buy",IF({PT}!AC{r}="",0,{PT}!AC{r})-ROW()*0.0000001,-9999-ROW()*0.0000001)'
    ws_dash[f"S{r}"] = f'=IF({PT}!BL{r}="Yes",IF({PT}!AC{r}="",0,{PT}!AC{r})-ROW()*0.0000001,-9999-ROW()*0.0000001)'
    ws_dash[f"T{r}"] = (
        f'=IF(OR({PT}!AY{r}="High",{PT}!BA{r}="High",AND({PT}!AB{r}<>"",{PT}!AB{r}<=2)),'
        f'IF({PT}!AC{r}="",0,{PT}!AC{r})-ROW()*0.0000001,-9999-ROW()*0.0000001)'
    )
for col in ("R", "S", "T"):
    ws_dash.column_dimensions[col].hidden = True

TABLE_COLS = [(1, 1), (2, 5), (6, 7), (8, 8), (9, 10), (11, 12)]  # Rank / Name / SKU / Score / Decision / Price


def section_title(row, text):
    cell = ws_dash[f"A{row}"]
    cell.value = text.upper()
    cell.font = Font(bold=True, size=11, color=DASH_WHITE)
    cell.alignment = Alignment(vertical="center", indent=1)
    for c in range(1, 13):
        ws_dash.cell(row=row, column=c).fill = PatternFill("solid", fgColor=DASH_ROSE)
    ws_dash.merge_cells(f"A{row}:L{row}")
    ws_dash.row_dimensions[row].height = 22


def table_header(row, labels):
    for (start, end), label in zip(TABLE_COLS, labels):
        sl, el = get_column_letter(start), get_column_letter(end)
        cell = ws_dash[f"{sl}{row}"]
        cell.value = label
        cell.font = Font(bold=True, color=DASH_ESPRESSO, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(start, end + 1):
            ws_dash.cell(row=row, column=c).fill = PatternFill("solid", fgColor=DASH_MIST)
            ws_dash.cell(row=row, column=c).border = header_bottom_border
        if end > start:
            ws_dash.merge_cells(f"{sl}{row}:{el}{row}")
    ws_dash.row_dimensions[row].height = 18


def kpi_card(start_row, slot, label, formula, accent_color, number_format=None, value_size=20):
    """One KPI card: small label row + a 2-row-tall big value, in a 3-column
    slot (4 slots fit across columns A-L), soft white background, thin warm
    beige border. Returns the value cell's address for later reuse."""
    col_start = 1 + slot * 3
    col_end = col_start + 2
    sl, el = get_column_letter(col_start), get_column_letter(col_end)
    label_row, value_row = start_row, start_row + 1

    ws_dash.merge_cells(f"{sl}{label_row}:{el}{label_row}")
    lc = ws_dash[f"{sl}{label_row}"]
    lc.value = label.upper()
    lc.font = Font(bold=True, size=8, color=DASH_TAUPE)
    lc.alignment = Alignment(horizontal="center", vertical="center")

    ws_dash.merge_cells(f"{sl}{value_row}:{el}{value_row + 1}")
    vc = ws_dash[f"{sl}{value_row}"]
    vc.value = formula
    vc.font = Font(bold=True, size=value_size, color=accent_color)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    if number_format:
        vc.number_format = number_format

    for r in range(label_row, value_row + 2):
        for c in range(col_start, col_end + 1):
            cell = ws_dash.cell(row=r, column=c)
            cell.fill = PatternFill("solid", fgColor=DASH_WHITE)
            cell.border = Border(
                left=beige_thin if c == col_start else None,
                right=beige_thin if c == col_end else None,
                top=beige_thin if r == label_row else None,
                bottom=beige_thin if r == value_row + 1 else None,
            )
    ws_dash.row_dimensions[label_row].height = 15
    ws_dash.row_dimensions[value_row].height = 20
    ws_dash.row_dimensions[value_row + 1].height = 20
    return f"{sl}{value_row}"


def apply_score_badges(rng):
    ws_dash.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["80"],
                         font=Font(color=GREEN_FONT, bold=True)))
    ws_dash.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["65", "79.999"],
                         font=Font(color=ORANGE_FONT, bold=True)))
    ws_dash.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["65"],
                         font=Font(color=RED_FONT, bold=True)))


def apply_decision_badges(rng):
    ws_dash.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Buy"'],
                         fill=PatternFill("solid", fgColor=GREEN), font=Font(color=GREEN_FONT, bold=True)))
    ws_dash.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Maybe"'],
                         fill=PatternFill("solid", fgColor=YELLOW), font=Font(color=YELLOW_FONT, bold=True)))
    ws_dash.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Price Review"'],
                         fill=PatternFill("solid", fgColor=ORANGE), font=Font(color=ORANGE_FONT, bold=True)))
    ws_dash.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Reject"'],
                         fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_FONT, bold=True)))


def add_topn_table(start_row, key_range, n):
    """n ranked rows pulled via LARGE()+MATCH() on a rank-key range (either a
    hidden Product Tracker column or a hidden Dashboard-local column above).
    Works in Excel/LibreOffice/Sheets alike (no QUERY/FILTER/SORT needed)."""
    band_fill = PatternFill("solid", fgColor=DASH_CREAM)
    for i in range(n):
        row = start_row + i
        key = f"LARGE({key_range},{i + 1})"
        match = f"MATCH({key},{key_range},0)"
        valid = f"{key}>-9000"
        ws_dash.cell(row=row, column=1, value=i + 1)
        ws_dash.merge_cells(f"B{row}:E{row}")
        ws_dash.cell(row=row, column=2,
                     value=f'=IF({valid},IFERROR(INDEX({PT}!$C${pt_first}:$C${pt_last},{match}),""),"")')
        ws_dash.merge_cells(f"F{row}:G{row}")
        ws_dash.cell(row=row, column=6,
                     value=f'=IF({valid},IFERROR(INDEX({PT}!$A${pt_first}:$A${pt_last},{match}),""),"")')
        ws_dash.cell(row=row, column=8,
                     value=f'=IF({valid},IFERROR(INDEX({PT}!$AC${pt_first}:$AC${pt_last},{match}),""),"")')
        ws_dash.merge_cells(f"I{row}:J{row}")
        ws_dash.cell(row=row, column=9,
                     value=f'=IF({valid},IFERROR(INDEX({PT}!$AD${pt_first}:$AD${pt_last},{match}),""),"")')
        ws_dash.merge_cells(f"K{row}:L{row}")
        ws_dash.cell(row=row, column=11,
                     value=f'=IF({valid},IFERROR(INDEX({PT}!$P${pt_first}:$P${pt_last},{match}),""),"")')
        for c in range(1, 13):
            cell = ws_dash.cell(row=row, column=c)
            cell.border = row_border
            cell.alignment = Alignment(horizontal="center" if c in (1, 8, 9, 11) else "left", vertical="center")
            if i % 2 == 1:
                cell.fill = band_fill
        ws_dash.cell(row=row, column=8).number_format = "0"
        ws_dash.cell(row=row, column=11).number_format = MONEY_FMT
        ws_dash.row_dimensions[row].height = 16
    apply_score_badges(f"H{start_row}:H{start_row + n - 1}")
    apply_decision_badges(f"I{start_row}:J{start_row + n - 1}")


# ---- Premium branded header ----
ws_dash.row_dimensions[1].height = 32
ws_dash.row_dimensions[2].height = 16
ws_dash.row_dimensions[3].height = 20
for c in range(1, 13):
    ws_dash.cell(row=1, column=c).fill = PatternFill("solid", fgColor=DASH_ROSE)
    ws_dash.cell(row=2, column=c).fill = PatternFill("solid", fgColor=DASH_ROSE)
    ws_dash.cell(row=3, column=c).fill = PatternFill("solid", fgColor=DASH_GOLD)
ws_dash.merge_cells("A1:L1")
ws_dash["A1"] = "AAYNA PRODUCT SCOUT"
ws_dash["A1"].font = Font(bold=True, size=20, color=DASH_WHITE)
ws_dash["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_dash.merge_cells("A2:L2")
ws_dash["A2"] = "Reflect your everyday style."
ws_dash["A2"].font = Font(italic=True, size=10, color=DASH_MIST)
ws_dash["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws_dash.merge_cells("A3:L3")
ws_dash["A3"] = "SOURCING DASHBOARD"
ws_dash["A3"].font = Font(bold=True, size=12, color=DASH_ESPRESSO)
ws_dash["A3"].alignment = Alignment(horizontal="center", vertical="center")

# ---- KPI Overview ----
section_title(5, "KPI Overview")
kpi_card(6, 0, "Total Products", f'=COUNTA({PT}!C{pt_first}:C{pt_last})', DASH_ESPRESSO)
kpi_card(6, 1, "Buy", f'=COUNTIF({PT}!AD{pt_first}:AD{pt_last},"Buy")', GREEN_FONT)
kpi_card(6, 2, "Maybe", f'=COUNTIF({PT}!AD{pt_first}:AD{pt_last},"Maybe")', YELLOW_FONT)
kpi_card(6, 3, "Price Review", f'=COUNTIF({PT}!AD{pt_first}:AD{pt_last},"Price Review")', ORANGE_FONT)
kpi_card(10, 0, "Reject", f'=COUNTIF({PT}!AD{pt_first}:AD{pt_last},"Reject")', RED_FONT)
kpi_card(10, 1, "Website Ready", f'=COUNTIF({PT}!BL{pt_first}:BL{pt_last},"Yes")', GREEN_FONT)
kpi_card(10, 2, "High Risk",
         f'=SUMPRODUCT(--(({PT}!AY{pt_first}:AY{pt_last}="High")+({PT}!BA{pt_first}:BA{pt_last}="High")'
         f'+(({PT}!AB{pt_first}:AB{pt_last}<>"")*({PT}!AB{pt_first}:AB{pt_last}<=2))>0))', RED_FONT)
addr_approved_cost = kpi_card(10, 3, "Approved Cost",
         f'=SUMIF({PT}!AG{pt_first}:AG{pt_last},"Approved",{PT}!AW{pt_first}:AW{pt_last})',
         DASH_GOLD, number_format=MONEY_FMT, value_size=16)

# ---- Budget Overview ----
section_title(14, "Budget Overview")
addr_cost = kpi_card(15, 0, "Approved Purchase Cost",
        f'=SUMIF({PT}!AG{pt_first}:AG{pt_last},"Approved",{PT}!AW{pt_first}:AW{pt_last})',
        DASH_ESPRESSO, number_format=MONEY_FMT, value_size=16)
addr_remaining = kpi_card(15, 1, "Remaining Budget", f"=MONTHLY_BUDGET-{addr_cost}",
        GREEN_FONT, number_format=MONEY_FMT, value_size=16)
kpi_card(15, 2, "Approved Quantity",
        f'=SUMIF({PT}!AG{pt_first}:AG{pt_last},"Approved",{PT}!AV{pt_first}:AV{pt_last})',
        DASH_TAUPE, number_format="#,##0", value_size=16)
kpi_card(15, 3, "Approved Products", f'=COUNTIF({PT}!AG{pt_first}:AG{pt_last},"Approved")',
        DASH_TAUPE, number_format="#,##0", value_size=16)

ws_dash.conditional_formatting.add(addr_remaining, CellIsRule(operator="lessThan", formula=["0"],
                     font=Font(color=RED_FONT, bold=True, size=16)))

# Budget Used progress bar
ws_dash.row_dimensions[18].height = 22
ws_dash.merge_cells("A18:C18")
ws_dash["A18"] = "BUDGET USED"
ws_dash["A18"].font = Font(bold=True, size=9, color=DASH_TAUPE)
ws_dash["A18"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_dash.merge_cells("D18:L18")
addr_budget_pct = "D18"
ws_dash[addr_budget_pct] = f'=IF(MONTHLY_BUDGET=0,"",{addr_cost}/MONTHLY_BUDGET)'
ws_dash[addr_budget_pct].number_format = "0%"
ws_dash[addr_budget_pct].font = Font(bold=True, size=12, color=DASH_ESPRESSO)
ws_dash[addr_budget_pct].alignment = Alignment(horizontal="center", vertical="center")
for c in range(1, 13):
    ws_dash.cell(row=18, column=c).border = row_border
ws_dash.conditional_formatting.add(addr_budget_pct, DataBarRule(
    start_type="num", start_value=0, end_type="num", end_value=1, color=DASH_GOLD))
ws_dash.conditional_formatting.add(addr_budget_pct, CellIsRule(operator="greaterThanOrEqual", formula=["1"],
                     font=Font(color=RED_FONT, bold=True, size=12)))
ws_dash.conditional_formatting.add(addr_budget_pct, CellIsRule(operator="greaterThanOrEqual", formula=["0.8"],
                     font=Font(color=ORANGE_FONT, bold=True, size=12)))

# ---- Top Products to Buy ----
section_title(20, "Top Products to Buy")
table_header(21, ["Rank", "Product Name", "SKU", "Score", "Decision", "Suggested Price"])
add_topn_table(22, f"$R${pt_first}:$R${pt_last}", 8)

# ---- Products Needing Partner Review ----
section_title(31, "Products Needing Partner Review")
table_header(32, ["Rank", "Product Name", "SKU", "Score", "Decision", "Suggested Price"])
add_topn_table(33, f"{PT}!$AN${pt_first}:$AN${pt_last}", 8)

# ---- Website Ready Products ----
section_title(42, "Website Ready Products")
table_header(43, ["Rank", "Product Name", "SKU", "Score", "Decision", "Suggested Price"])
add_topn_table(44, f"$S${pt_first}:$S${pt_last}", 6)

# ---- High Risk / Warning Products ----
section_title(51, "High Risk / Warning Products")
table_header(52, ["Rank", "Product Name", "SKU", "Score", "Decision", "Suggested Price"])
add_topn_table(53, f"$T${pt_first}:$T${pt_last}", 6)

# ---- Hidden chart-source data (kept off to the side, rows hidden) ----
ws_dash["A65"] = "Decision"
ws_dash["B65"] = "Count"
decision_labels = ["Buy", "Maybe", "Price Review", "Reject"]
decision_addrs = [None, None, None, None]
for i, lbl in enumerate(decision_labels):
    r = 66 + i
    ws_dash.cell(row=r, column=1, value=lbl)
    ws_dash.cell(row=r, column=2, value=f'=COUNTIF({PT}!AD{pt_first}:AD{pt_last},"{lbl}")')

ws_dash["A70"] = "Sourcing Status"
ws_dash["B70"] = "Count"
pipeline_stages = ["Not Started", "Sourcing", "Sample Ordered", "Sample Received",
                    "Approved - Not Ordered", "Ordered", "In Transit", "Arrived", "Live on Website"]
for i, stage in enumerate(pipeline_stages):
    r = 71 + i
    ws_dash.cell(row=r, column=1, value=stage)
    ws_dash.cell(row=r, column=2, value=f'=COUNTIF({PT}!BF{pt_first}:BF{pt_last},"{stage}")')

for r in range(65, 81):
    ws_dash.row_dimensions[r].hidden = True

# ---- Decision Breakdown chart ----
decision_chart = PieChart()
decision_chart.title = "Decision Breakdown"
decision_chart.height, decision_chart.width = 8, 12
data = Reference(ws_dash, min_col=2, min_row=65, max_row=69)
cats = Reference(ws_dash, min_col=1, min_row=66, max_row=69)
decision_chart.add_data(data, titles_from_data=True)
decision_chart.set_categories(cats)
decision_chart.series[0].data_points = [
    DataPoint(idx=0, spPr=GraphicalProperties(solidFill=GREEN_FONT)),
    DataPoint(idx=1, spPr=GraphicalProperties(solidFill=YELLOW_FONT)),
    DataPoint(idx=2, spPr=GraphicalProperties(solidFill=ORANGE_FONT)),
    DataPoint(idx=3, spPr=GraphicalProperties(solidFill=RED_FONT)),
]
ws_dash.add_chart(decision_chart, "N5")

# ---- Sourcing pipeline chart ----
pipeline_chart = BarChart()
pipeline_chart.type = "col"
pipeline_chart.title = "Sourcing Pipeline"
pipeline_chart.height, pipeline_chart.width = 8, 14
pdata = Reference(ws_dash, min_col=2, min_row=70, max_row=79)
pcats = Reference(ws_dash, min_col=1, min_row=71, max_row=79)
pipeline_chart.add_data(pdata, titles_from_data=True)
pipeline_chart.set_categories(pcats)
pipeline_chart.series[0].graphicalProperties = GraphicalProperties(solidFill=DASH_GOLD)
pipeline_chart.legend = None
ws_dash.add_chart(pipeline_chart, "N24")

# =====================================================================
# TAB: CLAUDE SCORING PROMPT
# =====================================================================
ws_prompt = wb.create_sheet("Claude Scoring Prompt")
ws_prompt.sheet_properties.tabColor = MAUVE
ws_prompt.column_dimensions["A"].width = 100

ws_prompt["A1"] = "AAYNA Product Scout Lite — Claude Scoring Prompt"
ws_prompt["A1"].font = Font(bold=True, size=14, color=ESPRESSO)

ws_prompt["A2"] = ("Copy everything in the box below into Claude (or another AI assistant), paste in the "
                    "product photo/link/description where indicated, then copy Claude's reply back into the "
                    "matching score columns on the Product Tracker tab.")
ws_prompt["A2"].font = note_font
ws_prompt["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws_prompt.row_dimensions[2].height = 40

claude_prompt = (
    "You are helping AAYNA, a Bangladesh-based women's accessories brand, evaluate a sourcing "
    "candidate product.\n\n"
    "AAYNA's brand identity: feminine, trendy, affordable, clean and modern, with a warm, "
    "social-media-friendly feel. Visual aesthetic: dusty rose, cream white, antique gold, and "
    "espresso charcoal. Most products sell under BDT 700.\n\n"
    "Here is the product: [paste product name, description, photos/link, and supplier price here]\n\n"
    "Score this product from 1 (worst) to 5 (best) on each of these 10 criteria, with a one-line "
    "reason for each score:\n\n"
    "1. Feminine Score - how feminine/delicate does it look?\n"
    "2. Trend Score - how on-trend is it right now (TikTok/Instagram/competitors)?\n"
    "3. AAYNA Aesthetic Fit Score - how well does it match our dusty rose / cream / antique gold look?\n"
    "4. Easy to Style Score - how easily can a customer style this with everyday outfits?\n"
    "5. Lightweight Score - how light/comfortable is it to wear or carry?\n"
    "6. Reels/Photo Potential Score - how good would this look in a Reel, photo, or unboxing video?\n"
    "7. Giftability Score - how good is this as a gift?\n"
    "8. Price Fit Score - how comfortably can we sell this under BDT 700 and still profit?\n"
    "9. Demand Score - how likely is our target customer to actually want and buy this?\n"
    "10. Quality/Risk Score - how low is the risk of poor quality, damage in transit, or returns?\n\n"
    "Then give:\n"
    "- A Recommended Decision: Buy, Maybe, Price Review, or Reject\n"
    "- A one-sentence Reason for that decision\n"
    "- A short Content/Reel Idea for how we'd promote this product\n\n"
    "Format your answer exactly like this so it's easy to paste into the sheet:\n"
    "Feminine: _ - reason\n"
    "Trend: _ - reason\n"
    "Aesthetic Fit: _ - reason\n"
    "Easy to Style: _ - reason\n"
    "Lightweight: _ - reason\n"
    "Reels/Photo: _ - reason\n"
    "Giftability: _ - reason\n"
    "Price Fit: _ - reason\n"
    "Demand: _ - reason\n"
    "Quality/Risk: _ - reason\n"
    "Recommended Decision: _\n"
    "Reason: _\n"
    "Content/Reel Idea: _"
)
ws_prompt["A4"] = claude_prompt
ws_prompt["A4"].alignment = Alignment(wrap_text=True, vertical="top")
ws_prompt["A4"].font = Font(name="Consolas", size=10, color=ESPRESSO)
ws_prompt["A4"].fill = PatternFill("solid", fgColor=CREAM)
thick_border = Border(left=Side(style="medium", color=ANTIQUE_GOLD), right=Side(style="medium", color=ANTIQUE_GOLD),
                       top=Side(style="medium", color=ANTIQUE_GOLD), bottom=Side(style="medium", color=ANTIQUE_GOLD))
ws_prompt["A4"].border = thick_border
ws_prompt.row_dimensions[4].height = 520

ws_prompt["A6"] = ("Tip: if you ever change the 10 scoring criteria, their order, or the price ceiling in the "
                    "Settings tab, update this prompt to match so Claude's answers keep lining up with the "
                    "Product Tracker columns.")
ws_prompt["A6"].font = note_font
ws_prompt["A6"].alignment = Alignment(wrap_text=True, vertical="top")
ws_prompt.row_dimensions[6].height = 40

# =====================================================================
# TAB: INSTRUCTIONS
# =====================================================================
ws_help = wb.create_sheet("Instructions")
ws_help.sheet_properties.tabColor = ESPRESSO
ws_help.column_dimensions["A"].width = 34
ws_help.column_dimensions["B"].width = 90

ws_help["A1"] = "AAYNA Product Scout Lite — How To Use"
ws_help["A1"].font = Font(bold=True, size=14, color=ESPRESSO)
ws_help.merge_cells("A1:B1")

intro = (
    "This sheet helps the 3 of us score and decide on products from SkyBuyBD, AliExpress, "
    "or other China sourcing platforms — fast, consistently, and without spreadsheets fights.\n\n"
    "Workflow: Whoever finds a product adds a new row and fills in the WHITE input columns only. "
    "The colored formula columns calculate themselves. Then each partner fills in the 10 score columns "
    "(or pastes in Claude's scores from the Claude Scoring Prompt tab), and the sheet tells you "
    "Buy / Maybe / Price Review / Reject automatically."
)
ws_help["A2"] = "Overview"
ws_help["A2"].font = label_font
ws_help["B2"] = intro
ws_help["B2"].alignment = Alignment(wrap_text=True, vertical="top")
ws_help.row_dimensions[2].height = 100

rows_help = [
    ("Step 1 — Add the product", "SKU fills in automatically. Enter Date Added, Product Name, Image/Link, "
        "Category, Source Platform, Source URL, Supplier Name."),
    ("Step 2 — Enter cost info", "Enter Unit Cost in whatever currency the supplier quoted (USD/RMB/BDT), pick the "
        "currency, then enter Shipping Cost/Unit and Misc Fees if known. Customs % and Misc Fees auto-fill from "
        "Settings tab if left blank."),
    ("Step 3 — Let it calculate", "Unit Cost (BDT), Estimated Landed Cost, Suggested Selling Price, Expected Profit, "
        "and Profit Margin % fill in automatically."),
    ("Step 4 — Score it (1 = worst, 5 = best)",
        "Feminine: how feminine/delicate does it look? "
        "Trend: how hot is this right now (TikTok/IG/competitors)? "
        "AAYNA Aesthetic Fit: how well does it match our dusty rose / cream / antique gold look? "
        "Easy to Style: how easily can a customer style this with everyday outfits? "
        "Lightweight: how light/comfortable is it to wear or carry? "
        "Reels/Photo Potential: how good would this look in a Reel, photo, or unboxing video? "
        "Giftability: how good is this as a gift? "
        "Price Fit: how comfortably can we sell this under BDT 700 and still profit? "
        "Demand: how likely is our target customer to actually want and buy this? "
        "Quality/Risk: how low is the risk of poor quality, damage in transit, or returns? "
        "Tip: use the Claude Scoring Prompt tab to get an AI first opinion on all 10 scores at once."),
    ("Step 5 — Read the verdict", "Total Product Score (0-100) calculates automatically using the weights set in "
        "the Settings tab. Decision applies these hard rules on top of the score: "
        "if Suggested Selling Price is over BDT 700, Decision becomes 'Price Review' (and 'Reject' if it's far "
        "over, above the Hard Reject Price Ceiling in Settings); if Quality/Risk or AAYNA Aesthetic Fit score is "
        "1 or 2, Decision can never be 'Buy' (it caps at 'Maybe') even if the total score is high. "
        "Buy = green, Maybe = yellow, Price Review = orange, Reject = red."),
    ("Step 6 — Decide together", "Add a short Reason so the other partners know WHY. Add a Content/Reel Idea so "
        "marketing has a head start. Set Approval Status once all partners agree, note who Approved and the date."),
    ("Step 7 — Sourcing & operations columns", "Once a product looks promising, fill in the sourcing/operations "
        "columns (after the scoring columns): Supplier Rating, Product Rating, Sold Count/Order Count, Review "
        "Count, and Real Review Photos? help you judge if a supplier is trustworthy. MOQ, Recommended Test "
        "Quantity, and Final Approved Quantity track how many units to order; Total Purchase Cost calculates "
        "itself (Estimated Landed Cost x Final Approved Quantity) once both are filled in. Weight/Size, "
        "Fragility Level, Packaging Difficulty, and Courier Risk flag shipping problems before you commit. "
        "Duplicate Found on BD Market?, Competitor Price, and Market Saturation Level help you judge if it's "
        "already oversaturated locally."),
    ("Auto Flag / Warning", "Calculates itself and shows the single biggest problem with a product, in priority "
        "order: 'Price too high' (over the price ceiling), 'Quality risk' (Quality/Risk score 1-2), 'Aesthetic "
        "mismatch' (Aesthetic Fit score 1-2), 'Too fragile' (Fragility Level = High), 'Low profit' (margin below "
        "the Settings threshold), 'MOQ too high' (above the Settings threshold), 'Needs partner review' (still "
        "Maybe/Price Review), or 'Good candidate' (a clean Buy with no problems). Use it as a quick at-a-glance "
        "warning, not a replacement for reading the Reason column."),
    ("Website-readiness checklist", "Product Name Finalized?, SKU Finalized?, Photos Ready?, Description Ready?, "
        "and Price Approved? are simple Yes/No checkboxes for whoever preps the listing. Ready for Website "
        "Upload? calculates itself: it only shows 'Yes' once ALL FIVE of those checklist columns say 'Yes'."),
    ("Sourcing Status", "Track where a product physically is in the pipeline: Not Started, Sourcing, Sample "
        "Ordered, Sample Received, Approved - Not Ordered, Ordered, In Transit, Arrived, or Live on Website. "
        "This is separate from Decision/Approval Status — a product can be Approved but still waiting on an "
        "order, or Ordered while the website checklist is still in progress."),
    ("How SKUs are generated", "SKU fills in automatically as AYN-<category code>-<number>, e.g. AYN-EAR-0001 "
        "for the first earrings row. The 3-letter code comes from a fixed lookup table (Earrings=EAR, "
        "Necklace=NEC, Ring=RNG, Bracelet=BRC, Hair Accessory=HAR, Gift Set=GFT, and one more for every other "
        "category) on the Dropdown Lists tab, not from the category spelling, so it never changes even if two "
        "categories start with the same letters."),
    ("Staying inside the monthly budget", "Set Monthly Inventory Budget on the Settings tab once. The Dashboard's "
        "Monthly Budget section then tracks Approved Purchase Cost (sum of Total Purchase Cost for every "
        "Approved product), Remaining Budget (budget minus that cost), Budget Used %, Approved Quantity, and "
        "Number of Approved Products — so you always know how much buying room is left before approving the "
        "next product."),
    ("Actual cost & profit (after ordering)", "SkyBuy/China sourcing costs can shift after shipping, fees, or "
        "exchange rate changes, so once an order actually ships, fill in Actual Product Cost and Actual Shipping "
        "Cost (both in BDT) and Actual Selling Price. Actual Landed Cost, Actual Profit, and Actual Profit "
        "Margin calculate themselves from those three, separately from the original Estimated Landed Cost/"
        "Suggested Selling Price — so you can see the real numbers without losing the original estimate."),
    ("Quality check after arrival", "Once a shipment arrives, fill in Arrival Date, Quantity Received, Defect "
        "Count, and QC Notes, then set QC Status: Not Arrived, QC Pending, QC Passed, Minor Defect, QC Failed, "
        "Discount Sell, or Return/Reject (color-coded green/yellow/orange/red). Final Stock Accepted calculates "
        "itself as Quantity Received minus Defect Count, so you know exactly how many good pieces you actually "
        "have to sell."),
    ("AI scoring trace", "Since Claude, ChatGPT, and Gemini can score the same product differently, record AI "
        "Tool Used, AI Score Date, Scored By, and Manual Score Adjusted? (Yes if a partner changed any AI-given "
        "score by hand) for every product you score with AI help. Over time this lets you compare which AI tool "
        "tends to line up best with the decisions the team actually makes."),
    ("Adjusting the formulas", "Don't like the default markup, exchange rate, price ceilings, scoring weights, "
        "low profit margin warning, MOQ warning threshold, or monthly budget? Change them ONCE in the Settings "
        "tab — every row updates automatically. No need to touch the Product Tracker formulas."),
    ("The Dashboard tab", "A live, read-only overview: how many products are Buy/Maybe/Price Review/Reject; "
        "total estimated purchase cost for approved products, products ready for website upload, high-risk "
        "products, products with high reel/photo potential, products approved but not ordered yet, and products "
        "ordered but not arrived yet; Monthly Budget tracking (Approved Purchase Cost, Remaining Budget, Budget "
        "Used %, Approved Quantity, Number of Approved Products); the Top 10 highest scoring products, the best "
        "products under BDT 700, the best products for reels/photos, the best giftable products, and everything "
        "still waiting on partner review. Nothing on this tab needs to be filled in by hand — it reads straight "
        "from the Product Tracker tab."),
    ("The Claude Scoring Prompt tab", "A ready-to-copy prompt for getting an AI first opinion on a product's 10 "
        "scores, a recommended decision, and a content/reel idea, formatted so the reply is easy to paste back "
        "into the right columns."),
    ("Color key", "Green = Buy / Approved. Yellow = Maybe / Pending. Orange = Price Review (price is over our 700 "
        "BDT target but might still be worth negotiating). Red = Reject / Rejected, or a Quality/Risk or "
        "Aesthetic Fit score of 1-2."),
    ("Tips for 3-partner teams", "Assign one score block per partner if you want speed (e.g. Partner A scores "
        "Feminine + Trend + Aesthetic Fit, Partner B scores Easy to Style + Lightweight + Reels/Photo, Partner C "
        "scores Giftability + Price Fit + Demand + Quality/Risk) — or all 3 score independently and average "
        "manually before finalizing. Keep the Reason column honest; it's your shared memory."),
    ("Where to enter new products", "Rows 3-7 are demo rows, one for each decision branch — leave them as "
        "examples or delete them once the team is comfortable. Start entering real products from row 8 onward; "
        "the Product Tracker has 498 ready-to-use rows (3-500). Never type over a colored, calculating column "
        "(SKU, costs, scores, Decision, Auto Flag/Warning, the readiness checklist, Actual Profit, or Final "
        "Stock Accepted) — only the white manual-input columns should be typed into."),
    ("Backups", "Keep one clean, untouched copy of this file somewhere safe before partners start editing, so "
        "there's always a working version to recover from if a formula ever gets accidentally overwritten."),
]
r = 4
for title, body in rows_help:
    ws_help[f"A{r}"] = title
    ws_help[f"A{r}"].font = label_font
    ws_help[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_help[f"B{r}"] = body
    ws_help[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_help.row_dimensions[r].height = 75
    r += 1

wb._sheets = [ws, ws_dash, ws_prompt, ws_set, ws_lists, ws_help]

out_path = "AAYNA_Product_Scout_Lite.xlsx"
wb.save(out_path)
print("Saved:", out_path)
